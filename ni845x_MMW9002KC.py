import ctypes
import numpy as np
import openpyxl as xl
from datetime import datetime
import time
import pyvisa

MAX_SIZE = 1024
DEV_SIZE = 256

ni845x_dll = ctypes.windll.LoadLibrary('Ni845x.dll')


class Ni845xError(Exception):
    def __init__(self, status_code):
        message = ctypes.create_string_buffer(MAX_SIZE)

        ni845x_dll.ni845xStatusToString(status_code, MAX_SIZE, message)

        Exception.__init__(self, message.value)


def errChk(err):
    if err:
        raise Ni845xError(err)


class NI845x():
    VOLTS33 = 33
    VOLTS25 = 25
    VOLTS18 = 18
    VOLTS15 = 15
    VOLTS12 = 12
    INPUT, OUTPUT = 0, 1
    kNi845xI2cAddress7Bit = 0
    kNi845xI2cAddress10Bit = 1

    def __init__(self):

        # Determine available devices
        NextDevice = ctypes.create_string_buffer(DEV_SIZE)
        FindDeviceHandle = ctypes.c_int32()
        NumberFound = ctypes.c_int32()

        errChk(ni845x_dll.ni845xFindDevice(ctypes.byref(NextDevice), ctypes.byref(FindDeviceHandle),
                                           ctypes.byref(NumberFound)))

        if NumberFound.value != 1:
            raise Exception('Only implemented support for exactly 1 USB card. {} found.'.format(NumberFound.value))
        self._name = NextDevice
        self._open()
        self.config_i2c()
        self.set_io_voltage_level(self.VOLTS18)  # should never exceed 1.8 for MMW9002 Beamstearing device
        self.set_port_line_direction_map(self.OUTPUT * np.ones(8))  # all lines set to be output

    def Script_I2C(self, Clock_Rate=100):
        ScriptHandle = ctypes.c_uint32()
        clock_rate = ctypes.c_uint16(Clock_Rate)
        errChk(ni845x_dll.ni845xI2cScriptOpen(ctypes.byref(ScriptHandle)))
        errChk(ni845x_dll.ni845xI2cScriptClockRate(ScriptHandle, clock_rate))
        return ScriptHandle

    def Script_I2C_Read(self, RegisterNumber=0x06, NumberToRead=2, Address=0x2B, PortNumber=0):
        ScriptHandle = self.Script_I2C()
        Address = ctypes.c_uint8(Address)
        WriteData = ctypes.create_string_buffer(b'', 1)
        NumberToRead = ctypes.c_uint16(NumberToRead)
        PortNumber = ctypes.c_uint8(PortNumber)
        WriteDataSize = ctypes.c_uint32(1)
        WriteData[0] = RegisterNumber
        ScriptReadIndex = ctypes.c_uint32()
        ReadDataSize = ctypes.c_uint32(0)
        ReadData = ctypes.create_string_buffer(b'', NumberToRead.value)
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, WriteDataSize, WriteData))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressRead(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptRead(ScriptHandle, NumberToRead, ctypes.c_uint32(1),
                                              ctypes.byref(ScriptReadIndex)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStop(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cRunScript(ScriptHandle, self.dev_handle, PortNumber))
        errChk(ni845x_dll.ni845xI2cScriptExtractReadDataSize(ScriptHandle, ScriptReadIndex, ctypes.byref(ReadDataSize)))
        errChk(ni845x_dll.ni845xI2cScriptExtractReadData(ScriptHandle, ScriptReadIndex, ctypes.byref(ReadData)))
        return ReadData

    def Script_I2C_Write(self, RegisterNumber=0x06, Data=b'0', Address=0x2B, PortNumber=0):
        ScriptHandle = self.Script_I2C()
        Address = ctypes.c_uint8(Address)
        WriteData = ctypes.create_string_buffer(b'', len(Data) + 1)
        PortNumber = ctypes.c_uint8(PortNumber)
        WriteDataSize = ctypes.c_uint32(len(Data) + 1)
        WriteData[0] = RegisterNumber
        for i in range(len(Data)):
            WriteData[i + 1] = Data[i]
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, WriteDataSize, WriteData))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStop(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cRunScript(ScriptHandle, self.dev_handle, PortNumber))


    def Script_Write_LUTValue(self, Data= 4, LUTRow= 0,  Channel=0, mode="TX", property ="gain",chk =0, Address=0x2B, PortNumber=0,):
        if(Data != Add_Parity_Bit(Data)):
            chk +=1
            if(LUTRow ==0):
                LUTRow=63
            else:
                LUTRow-=1
            PrevValue = self.Script_Read_LUTValue(1, LUTRow, Channel, mode, property)
        else:
            chk = 0
        Data = bytes(chr(Add_Parity_Bit(Data)),'utf8')
        ScriptHandle = self.Script_I2C()
        Address = ctypes.c_uint8(Address)
        WriteData = ctypes.create_string_buffer(b'', len(Data) + 1)
        PortNumber = ctypes.c_uint8(PortNumber)
        WriteDataSize = ctypes.c_uint32(len(Data) + 1)
        WriteData[0] = 0x02
        for i in range(len(Data)):
            WriteData[i + 1] = Data[i]
        Reg0Value = ctypes.create_string_buffer(b'', 2)
        Reg0Value[0] = 0x00
        Reg0Value[1] = Add_Parity_Bit(LUTRow)           #very first address of look up table
        Reg1Value = ctypes.create_string_buffer(b'', 2)
        Reg1Value[0] = 0x01
        Reg1Value[1] = bytes(chr(self.getReg1Value(Channel,mode,property)),'utf8')[0]
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, ctypes.c_uint32(2), Reg0Value))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, ctypes.c_uint32(2), Reg1Value))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, WriteDataSize, WriteData))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStop(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cRunScript(ScriptHandle, self.dev_handle, PortNumber))
        if(chk < 63 and chk > 0):
            self.Script_Write_LUTValue(ord(PrevValue[0]), LUTRow, Channel, mode, property, chk)


    def Script_Read_LUTValue(self, nbytes = 1,LUTRow=0x80,  Channel=0, mode="TX", property ="gain", Address=0x2B, PortNumber=0 ):
        ScriptHandle = self.Script_I2C()
        Address = ctypes.c_uint8(Address)
        PortNumber = ctypes.c_uint8(PortNumber)
        ScriptReadIndex = ctypes.c_uint32()
        Reg2Value = ctypes.create_string_buffer(b'', 1)
        Reg2Value[0] = 0x02
        Reg0Value = ctypes.create_string_buffer(b'', 2)
        Reg0Value[0] = 0x00
        Reg0Value[1] = Add_Parity_Bit(LUTRow)           #very first address of look up table
        Reg1Value = ctypes.create_string_buffer(b'', 2)
        Reg1Value[0] = 0x01
        Reg1Value[1] = bytes(chr(self.getReg1Value(Channel,mode,property)),'utf8')[0]
        nbytes = ctypes.c_uint16(nbytes)
        ReadData = ctypes.create_string_buffer(b'', nbytes.value)
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, ctypes.c_uint32(2), Reg0Value))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, ctypes.c_uint32(2), Reg1Value))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressWrite(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptWrite(ScriptHandle, ctypes.c_uint32(1), Reg2Value))
        errChk(ni845x_dll.ni845xI2cScriptDelay(ScriptHandle, ctypes.c_uint8(100)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStart(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cScriptAddressRead(ScriptHandle, Address))
        errChk(ni845x_dll.ni845xI2cScriptRead(ScriptHandle, nbytes, ctypes.c_uint32(1),
                                              ctypes.byref(ScriptReadIndex)))
        errChk(ni845x_dll.ni845xI2cScriptIssueStop(ScriptHandle))
        errChk(ni845x_dll.ni845xI2cRunScript(ScriptHandle, self.dev_handle, PortNumber))
        errChk(ni845x_dll.ni845xI2cScriptExtractReadData(ScriptHandle, ScriptReadIndex, ctypes.byref(ReadData)))
        return ReadData


    def _open(self):
        self.dev_handle = ctypes.c_int32()
        errChk(ni845x_dll.ni845xOpen(ctypes.byref(self._name), ctypes.byref(self.dev_handle)))

    def set_port_line_direction_map(self, mapp, port=0):
        # mapp: np array or list with 8 0's or 1's
        # 0 = input, 1 = output
        port = ctypes.c_uint8(port)
        mapp = np.asarray(mapp)
        print(mapp)
        assert len(mapp) == 8
        r = np.arange(7, -1, -1)
        _map = np.sum(2 ** r * mapp).astype(int)
        bitmap = ctypes.c_uint8(_map)
        errChk(ni845x_dll.ni845xDioSetPortLineDirectionMap(self.dev_handle, port, bitmap))

    def set_io_voltage_level(self, lev):
        lev = ctypes.c_uint8(lev)
        errChk(ni845x_dll.ni845xSetIoVoltageLevel(self.dev_handle, lev))

    def end(self):
        errChk(ni845x_dll.ni845xClose(self.dev_handle))
        errChk(ni845x_dll.ni845xI2cConfigurationClose(self.i2c_handle))

    def write_dio(self, line, val, port=0):
        line = ctypes.c_uint8(line)
        port = ctypes.c_uint8(port)
        val = ctypes.c_int32(val)
        output = errChk(ni845x_dll.ni845xDioWriteLine(self.dev_handle, port, line, val))
        print(output)

    def config_i2c(self, size=None, address=0x2B, clock_rate=100, timeout=20000, pullupenable=1):
        """
        Set the ni845x I2C configuration.

        Parameters
        ----------
            size : Configuration address size (default 7Bit).
            address : Configuration address (default 0).
            clock_rate : Configuration clock rate in kilohertz (default 100).

        Returns
        -------
            None
        """
        if size is None:
            size = self.kNi845xI2cAddress7Bit

        size = ctypes.c_int32(size)
        address = ctypes.c_uint16(address)
        clock_rate = ctypes.c_uint16(clock_rate)
        timeout = ctypes.c_uint32(timeout)
        pullupenable = ctypes.c_uint8(pullupenable)
        #
        # create configuration reference
        #
        self.i2c_handle = ctypes.c_int32()
        errChk(ni845x_dll.ni845xI2cSetPullupEnable(self.dev_handle, pullupenable))
        errChk(ni845x_dll.ni845xI2cConfigurationOpen(ctypes.byref(self.i2c_handle)))

        #
        # configure configuration properties
        #
        errChk(ni845x_dll.ni845xI2cConfigurationSetAddressSize(self.i2c_handle, size))
        errChk(ni845x_dll.ni845xI2cConfigurationSetAddress(self.i2c_handle, address))
        errChk(ni845x_dll.ni845xI2cConfigurationSetClockRate(self.i2c_handle, clock_rate))
        errChk(ni845x_dll.ni845xSetTimeout(self.dev_handle, timeout))
        print("Configuration Completed")

    def write_i2c(self, data):
        """
        Write an array of data to an I2C slave device.

        Parameters
        ----------
            write_data : Array of bytes to be written. Should be convertible to numpy array of
                    type unsignepip d char.

        Returns
        -------
            None

        """
        nbytes = ctypes.c_int32(len(data))
        data = ctypes.addressof(ctypes.create_string_buffer(data, len(data)))
        x = errChk(ni845x_dll.ni845xI2cWrite(self.dev_handle, self.i2c_handle, nbytes, data))




    def Read_LUT(self):
        my_wb = xl.Workbook()
        my_sheet = my_wb.active
        Column = 0
        Modes = ['RX','TX']
        property = ['PHASE','GAIN']
        for channel in range(4):
            for mode in Modes:
                for pr in property:
                    Column +=1
                    Data = self.Script_Read_LUTValue(64, 0, channel, mode, pr)
                    for i in range(64):
                        c1 = my_sheet.cell(row=3 + i, column = Column)
                        c1.value = ord(Data[i])
        my_wb.save("Read_LUT-" + datetime.now().strftime("%d-%m-%Y,%H-%M") + ".xlsx")



    def Write_LUT(self, input_LUT = "input_LUT.xlsx"):
        my_wb = xl.load_workbook(input_LUT)
        my_sheet = my_wb.active
        Column = 0
        Modes = ['RX', 'TX']
        property = ['PHASE', 'GAIN']
        for channel in range(4):
            for mode in Modes:
                for pr in property:
                    Column += 1
                    for i in range(64):
                        self.Script_Write_LUTValue(int(my_sheet.cell(row=3 + i, column=Column).value), i, channel, mode, pr)


    def getReg1Value(self, Channel, mode, property):
        if (mode == "TX" or mode == "tx"):
            if (property == "gain" or property == "GAIN"):
                reg1 = 0x30
            elif (property == "phase" or property == "PHASE"):
                reg1 = 0x20
            else:
                print("invalid property value")
        elif (mode == "RX" or mode == "rx"):
            if (property == "gain" or property == "GAIN"):
                reg1 = 0x10
            elif (property == "phase" or property == "PHASE"):
                reg1 = 0x00
            else:
                print("invalid property value")
        else:
            print("invalid mode value")
        reg1 = reg1 | 64 * Channel
        return reg1


    def Channel_setProperty(self, Channel = 0, mode = "tx", property = "gain" ,value = 0 ):
        reg1 = self.getReg1Value(Channel, mode, property)
        self.Script_I2C_Write(0x00, bytes(chr(Add_Parity_Bit(2)), 'utf8'))
        time.sleep(.1)
        self.Script_I2C_Write(0x01, bytes(chr(reg1), 'utf8'))
        time.sleep(.1)
        #self.Script_I2C_Write(0x02, bytes(chr(Add_Parity_Bit(value)), 'utf8'))
        #time.sleep(.1)
        return self.Script_I2C_Read(0x02)[0]


def main():
    Device = NI845x()
    time.sleep(1)
    Device.Script_I2C_Write(0x06, bytes(chr(0x44), 'utf8'))
    Data = Device.Script_I2C_Read(0x03)
    print(Data[0])
    rm = pyvisa.ResourceManager()
    rs = rm.list_resources()
    print(rs)
    PowerSupplyAddress = str(rm.list_resources()[0])  #if you have power supply connected to computer otherwise remove the lines below in main function
    PowerSupply = rm.open_resource(PowerSupplyAddress)
    Device.Read_LUT()
    for i in range(8):
        Device.Script_I2C_Write(0x06, bytes(chr(i*16+i), 'utf8'))
        print(float(PowerSupply.query('MEAS:CURR? (@2)')))
    Device.end()


def Add_Parity_Bit(input):
    count = 0
    input = input & 0x3F
    output = input
    while (input):
        count += input & 1
        input >>= 1
    if (count % 2 == 0):
        return output | 0x80
    else:
        return output


def Remove_Parity_Bit(input):
    return input & 0x3F


if __name__ == '__main__':
    main()
